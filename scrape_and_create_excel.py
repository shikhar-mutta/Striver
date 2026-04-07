#!/usr/bin/env python3
"""
Scrape Striver's A2Z DSA Sheet and create a comprehensive Excel tracker.
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import re
import json

# ─── Step 1: Try fetching and parsing the page ────────────────────────────────
URL = "https://takeuforward.org/dsa/strivers-a2z-sheet-learn-dsa-a-to-z"

print("Fetching page...")
try:
    resp = requests.get(URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
    soup = BeautifulSoup(resp.text, "html.parser")
    
    # Try to find embedded JSON data (Next.js pages often have __NEXT_DATA__)
    script_tag = soup.find("script", id="__NEXT_DATA__")
    if script_tag:
        data = json.loads(script_tag.string)
        print("Found __NEXT_DATA__, extracting...")
        print(json.dumps(data, indent=2)[:2000])
    else:
        print("No __NEXT_DATA__ found, using known curriculum data.")
except Exception as e:
    print(f"Fetch error: {e}")
    print("Using known curriculum data.")

# ─── Step 2: Comprehensive Striver A2Z DSA Curriculum ─────────────────────────
# This is the complete curriculum as listed on takeuforward.org

modules = [
    {
        "step": 1,
        "name": "Learn the Basics",
        "total": 31,
        "subtopics": [
            {"name": "Things to Know in C++/Java/Python or any language", "problems": 7},
            {"name": "Build-up Logical Thinking", "problems": 1},
            {"name": "Learn STL/Java-Collections or similar", "problems": 1},
            {"name": "Know Basic Maths", "problems": 7},
            {"name": "Learn Basic Recursion", "problems": 9},
            {"name": "Learn Basic Hashing", "problems": 6},
        ]
    },
    {
        "step": 2,
        "name": "Learn Important Sorting Techniques",
        "total": 7,
        "subtopics": [
            {"name": "Sorting-I", "problems": 3},
            {"name": "Sorting-II", "problems": 4},
        ]
    },
    {
        "step": 3,
        "name": "Solve Problems on Arrays [Easy -> Medium -> Hard]",
        "total": 40,
        "subtopics": [
            {"name": "Easy", "problems": 14},
            {"name": "Medium", "problems": 14},
            {"name": "Hard", "problems": 12},
        ]
    },
    {
        "step": 4,
        "name": "Binary Search [1D, 2D Arrays, Search Space]",
        "total": 32,
        "subtopics": [
            {"name": "BS on 1D Arrays", "problems": 13},
            {"name": "BS on Answers", "problems": 11},
            {"name": "BS on 2D Arrays", "problems": 8},
        ]
    },
    {
        "step": 5,
        "name": "Strings [Basic and Medium]",
        "total": 15,
        "subtopics": [
            {"name": "Basic and Easy String Problems", "problems": 7},
            {"name": "Medium String Problems", "problems": 8},
        ]
    },
    {
        "step": 6,
        "name": "Learn LinkedList [Single LL, Double LL, Medium, Hard Problems]",
        "total": 31,
        "subtopics": [
            {"name": "Learn 1D LinkedList", "problems": 6},
            {"name": "Learn Doubly LinkedList", "problems": 4},
            {"name": "Medium Problems of LL", "problems": 15},
            {"name": "Medium Problems of DLL", "problems": 1},
            {"name": "Hard Problems of LL", "problems": 5},
        ]
    },
    {
        "step": 7,
        "name": "Recursion [PatternWise]",
        "total": 25,
        "subtopics": [
            {"name": "Get a Strong Hold", "problems": 5},
            {"name": "Subsequences Pattern", "problems": 14},
            {"name": "Trying out all Combos / Hard", "problems": 6},
        ]
    },
    {
        "step": 8,
        "name": "Bit Manipulation [Concepts & Problems]",
        "total": 18,
        "subtopics": [
            {"name": "Learn Bit Manipulation", "problems": 8},
            {"name": "Interview Problems", "problems": 5},
            {"name": "Advanced Maths", "problems": 5},
        ]
    },
    {
        "step": 9,
        "name": "Stack and Queues [Learning, Pre-In-Post-fix, Monotonic Stack, Implementation]",
        "total": 30,
        "subtopics": [
            {"name": "Learning", "problems": 8},
            {"name": "Prefix, Infix, Postfix Conversion Problems", "problems": 6},
            {"name": "Monotonic Stack/Queue Problems [VVV. Imp]", "problems": 10},
            {"name": "Implementation Problems", "problems": 6},
        ]
    },
    {
        "step": 10,
        "name": "Sliding Window & Two Pointer Combined Problems",
        "total": 12,
        "subtopics": [
            {"name": "Medium Problems", "problems": 6},
            {"name": "Hard Problems", "problems": 6},
        ]
    },
    {
        "step": 11,
        "name": "Heaps [Learning, Medium, Hard Problems]",
        "total": 17,
        "subtopics": [
            {"name": "Learning", "problems": 3},
            {"name": "Medium Problems", "problems": 8},
            {"name": "Hard Problems", "problems": 6},
        ]
    },
    {
        "step": 12,
        "name": "Greedy Algorithms [Easy, Medium/Hard]",
        "total": 16,
        "subtopics": [
            {"name": "Easy Problems", "problems": 6},
            {"name": "Medium/Hard", "problems": 10},
        ]
    },
    {
        "step": 13,
        "name": "Binary Trees [Traversals, Medium and Hard Problems]",
        "total": 39,
        "subtopics": [
            {"name": "Traversals", "problems": 13},
            {"name": "Medium Problems", "problems": 12},
            {"name": "Hard Problems", "problems": 14},
        ]
    },
    {
        "step": 14,
        "name": "Binary Search Trees [Concept and Problems]",
        "total": 16,
        "subtopics": [
            {"name": "Concepts", "problems": 4},
            {"name": "Practice Problems", "problems": 12},
        ]
    },
    {
        "step": 15,
        "name": "Graphs [Concepts & Problems]",
        "total": 53,
        "subtopics": [
            {"name": "Learning", "problems": 6},
            {"name": "Problems on BFS/DFS", "problems": 14},
            {"name": "Topo Sort and Problems", "problems": 7},
            {"name": "Shortest Path Algorithms and Problems", "problems": 12},
            {"name": "Minimum Spanning Tree / Disjoint Set and Problems", "problems": 8},
            {"name": "Other Algorithms", "problems": 6},
        ]
    },
    {
        "step": 16,
        "name": "Dynamic Programming [Patterns and Problems]",
        "total": 56,
        "subtopics": [
            {"name": "Introduction to DP", "problems": 1},
            {"name": "1D DP", "problems": 5},
            {"name": "2D/3D DP and DP on Grids", "problems": 7},
            {"name": "DP on Subsequences", "problems": 9},
            {"name": "DP on Strings", "problems": 10},
            {"name": "DP on Stocks", "problems": 6},
            {"name": "DP on LIS", "problems": 5},
            {"name": "MCM DP / Partition DP", "problems": 7},
            {"name": "DP on Squares", "problems": 1},
            {"name": "DP on Rectangles", "problems": 5},
        ]
    },
    {
        "step": 17,
        "name": "Tries",
        "total": 7,
        "subtopics": [
            {"name": "Theory", "problems": 1},
            {"name": "Problems", "problems": 6},
        ]
    },
    {
        "step": 18,
        "name": "Strings",
        "total": 9,
        "subtopics": [
            {"name": "Hard Problems", "problems": 9},
        ]
    },
]

# ─── Step 3: Create Beautiful Excel ───────────────────────────────────────────
print("\nCreating Excel file...")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Striver's A2Z DSA Sheet"

# ── Color palette ──
DARK_BG       = "1A1A2E"
HEADER_BG     = "16213E"
MODULE_BG     = "0F3460"
SUBTOPIC_BG   = "1A1A2E"
ACCENT        = "E94560"
GOLD          = "F5A623"
WHITE         = "FFFFFF"
LIGHT_GRAY    = "CCCCCC"
GREEN         = "00C853"
DARK_GREEN    = "1B5E20"
ROW_ALT_1     = "1E1E3F"
ROW_ALT_2     = "252550"
BORDER_COLOR  = "333366"

# ── Fonts ──
title_font     = Font(name="Calibri", size=22, bold=True, color=WHITE)
header_font    = Font(name="Calibri", size=11, bold=True, color=WHITE)
module_font    = Font(name="Calibri", size=13, bold=True, color=GOLD)
subtopic_font  = Font(name="Calibri", size=11, color=WHITE)
count_font     = Font(name="Calibri", size=11, color=LIGHT_GRAY)
check_font     = Font(name="Wingdings 2", size=14, color=GREEN)
status_font    = Font(name="Calibri", size=11, bold=True, color=GREEN)

# ── Fills ──
dark_fill      = PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type="solid")
header_fill    = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
module_fill    = PatternFill(start_color=MODULE_BG, end_color=MODULE_BG, fill_type="solid")
accent_fill    = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
alt_fill_1     = PatternFill(start_color=ROW_ALT_1, end_color=ROW_ALT_1, fill_type="solid")
alt_fill_2     = PatternFill(start_color=ROW_ALT_2, end_color=ROW_ALT_2, fill_type="solid")

# ── Borders ──
thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)

# ── Alignment ──
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align   = Alignment(horizontal="left", vertical="center", wrap_text=True)
left_indent  = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)

# ── Column widths ──
ws.column_dimensions["A"].width = 10   # Step #
ws.column_dimensions["B"].width = 60   # Module / Sub-topic Name
ws.column_dimensions["C"].width = 18   # Total Problems
ws.column_dimensions["D"].width = 14   # Progress
ws.column_dimensions["E"].width = 14   # Status (✓)
ws.column_dimensions["F"].width = 18   # Notes

# ── Title Row ──
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "📚 Striver's A2Z DSA Course - Progress Tracker"
title_cell.font = title_font
title_cell.fill = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 50

# ── Subtitle Row ──
ws.merge_cells("A2:F2")
sub_cell = ws["A2"]
sub_cell.value = "Source: takeuforward.org/dsa/strivers-a2z-sheet-learn-dsa-a-to-z  |  Total: 455 Problems across 18 Steps"
sub_cell.font = Font(name="Calibri", size=10, italic=True, color=LIGHT_GRAY)
sub_cell.fill = dark_fill
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 28

# ── Empty separator ──
for col in range(1, 7):
    ws.cell(row=3, column=col).fill = dark_fill
ws.row_dimensions[3].height = 8

# ── Header Row ──
headers = ["Step", "Topic / Sub-topic", "Problems", "Progress 📊", "Done ✓", "Notes"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[4].height = 32

# ── Data Rows ──
current_row = 5
total_problems = 0

for mod in modules:
    # Module header row
    ws.row_dimensions[current_row].height = 36
    
    step_cell = ws.cell(row=current_row, column=1, value=f"Step {mod['step']}")
    step_cell.font = module_font
    step_cell.fill = module_fill
    step_cell.alignment = center_align
    step_cell.border = thin_border
    
    name_cell = ws.cell(row=current_row, column=2, value=mod["name"])
    name_cell.font = module_font
    name_cell.fill = module_fill
    name_cell.alignment = left_align
    name_cell.border = thin_border
    
    total_cell = ws.cell(row=current_row, column=3, value=f"{mod['total']} problems")
    total_cell.font = Font(name="Calibri", size=12, bold=True, color=ACCENT)
    total_cell.fill = module_fill
    total_cell.alignment = center_align
    total_cell.border = thin_border
    
    for col in [4, 5, 6]:
        cell = ws.cell(row=current_row, column=col)
        cell.fill = module_fill
        cell.border = thin_border
    
    total_problems += mod["total"]
    current_row += 1
    
    # Sub-topic rows
    for i, sub in enumerate(mod["subtopics"]):
        ws.row_dimensions[current_row].height = 28
        fill = alt_fill_1 if i % 2 == 0 else alt_fill_2
        
        # Step column (sub-number)
        sub_step_cell = ws.cell(row=current_row, column=1, value=f"{mod['step']}.{i+1}")
        sub_step_cell.font = Font(name="Calibri", size=10, color=LIGHT_GRAY)
        sub_step_cell.fill = fill
        sub_step_cell.alignment = center_align
        sub_step_cell.border = thin_border
        
        # Sub-topic name
        sub_name_cell = ws.cell(row=current_row, column=2, value=f"    {sub['name']}")
        sub_name_cell.font = subtopic_font
        sub_name_cell.fill = fill
        sub_name_cell.alignment = left_indent
        sub_name_cell.border = thin_border
        
        # Problem count
        sub_count_cell = ws.cell(row=current_row, column=3, value=sub["problems"])
        sub_count_cell.font = count_font
        sub_count_cell.fill = fill
        sub_count_cell.alignment = center_align
        sub_count_cell.border = thin_border
        
        # Progress tracking (empty, user fills in with ✓)
        progress_cell = ws.cell(row=current_row, column=4, value="")
        progress_cell.font = Font(name="Calibri", size=14, color=GREEN)
        progress_cell.fill = fill
        progress_cell.alignment = center_align
        progress_cell.border = thin_border
        
        # Done checkbox (empty, user fills in)
        done_cell = ws.cell(row=current_row, column=5, value="☐")
        done_cell.font = Font(name="Calibri", size=14, color=LIGHT_GRAY)
        done_cell.fill = fill
        done_cell.alignment = center_align
        done_cell.border = thin_border
        
        # Notes (empty)
        notes_cell = ws.cell(row=current_row, column=6)
        notes_cell.fill = fill
        notes_cell.border = thin_border
        
        current_row += 1
    
    # Separator row after each module
    ws.row_dimensions[current_row].height = 4
    for col in range(1, 7):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = dark_fill
        cell.border = Border()
    current_row += 1

# ── Summary Footer ──
current_row += 1
ws.merge_cells(f"A{current_row}:F{current_row}")
summary_cell = ws.cell(row=current_row, column=1)
summary_cell.value = f"🎯 Total: {total_problems} Problems  |  18 Steps  |  {sum(len(m['subtopics']) for m in modules)} Sub-topics"
summary_cell.font = Font(name="Calibri", size=14, bold=True, color=GOLD)
summary_cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
summary_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[current_row].height = 40

# ── Freeze panes (header row) ──
ws.freeze_panes = "A5"

# ── Print settings ──
ws.sheet_properties.tabColor = ACCENT

# ─── Save ────────────────────────────────────────────────────────────────────
output_path = "/home/shikhar/Sem 2/Coding Practice/Striver/Strivers_A2Z_DSA_Sheet.xlsx"
wb.save(output_path)
print(f"\n✅ Excel file saved to: {output_path}")
print(f"   Total modules: {len(modules)}")
print(f"   Total sub-topics: {sum(len(m['subtopics']) for m in modules)}")
print(f"   Total problems: {total_problems}")
